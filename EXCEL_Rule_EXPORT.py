import openpyxl
from help import RemoveWhiteSpaces


def IMPORT_RULES_FROM_EXCEL():
    dict = {}
    wb = openpyxl.load_workbook('The_tester_table_Version_6.xlsx')
    sheet = wb.active

    for i in range(2, 24):

        non_terminal = sheet.cell(i,1).value
        non_terminal = str(non_terminal)
        non_terminal = non_terminal.strip()
        non_terminal = RemoveWhiteSpaces(non_terminal)

        for num in range(2,33):

            terminal = sheet.cell(1,num).value
            terminal = str(terminal)
            terminal = terminal.strip()
            terminal = RemoveWhiteSpaces(terminal)

            slot_intersection_of_terminal_and_non_terminal = sheet.cell(i,num).value
            slot_intersection_of_terminal_and_non_terminal = str(slot_intersection_of_terminal_and_non_terminal)
            slot_intersection_of_terminal_and_non_terminal = slot_intersection_of_terminal_and_non_terminal.strip()
            slot_intersection_of_terminal_and_non_terminal = RemoveWhiteSpaces(slot_intersection_of_terminal_and_non_terminal)

            if slot_intersection_of_terminal_and_non_terminal == 'None':
                slot_intersection_of_terminal_and_non_terminal = ''

            if num == 2:
                dict[non_terminal] = {terminal: slot_intersection_of_terminal_and_non_terminal}
            else:
                dict[non_terminal][terminal] = slot_intersection_of_terminal_and_non_terminal

    return dict
